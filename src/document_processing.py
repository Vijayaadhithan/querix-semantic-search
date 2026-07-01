import csv
import hashlib
import json
import logging
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from pypdf import PdfReader

from database_store import (
    DatabaseRuntimeConfig,
    database_backend,
    database_source_name,
)
from settings import (
    CHUNK_OVERLAP,
    CHUNK_SIZE,
    EMBED_MODEL,
    MYSQL_DATABASE,
    MYSQL_SEARCH_ID_COLUMN,
    MYSQL_TABLE,
    SOURCE_FILES_DIR,
)

MYSQL_METADATA_MAX_CHARS = 1000
MYSQL_JSON_METADATA_MAX_FIELDS = 64
SUPPORTED_EXTENSIONS = {".pdf", ".csv", ".tsv", ".xlsx", ".xlsm"}
JSON_TEXT_KEYS = (
    "embedding_text",
    "semantic_text",
    "search_text",
    "text",
    "content",
    "document",
)
LABELED_TEXT_KEYS = (
    "Title",
    "Description",
    "Listing meta title",
    "Listing meta description",
    "Main category",
    "Main category meta title",
    "Main category meta description",
    "Subcategory",
    "Subcategory meta title",
    "Subcategory meta description",
    "Listing rental duration",
    "State",
    "City",
    "Locality",
    "Selected attributes",
    "Selected attribute values",
)
LABELED_TEXT_PATTERN = re.compile(
    r"(?<!^)\s+("
    + "|".join(
        re.escape(label)
        for label in sorted(LABELED_TEXT_KEYS, key=len, reverse=True)
    )
    + r"):\s*"
)


class _PypdfRepairFilter(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        return not record.getMessage().startswith("Ignoring wrong pointing object")


logging.getLogger("pypdf._reader").addFilter(_PypdfRepairFilter())


def read_pdf(path: Path) -> tuple[list[dict], int, int]:
    reader = PdfReader(path)
    pages = []
    empty_pages = 0
    for page_number, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        if text:
            pages.append({"page": page_number, "text": text})
        else:
            empty_pages += 1
    return pages, empty_pages, len(reader.pages)


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_headers(values: list[Any]) -> list[str]:
    headers = []
    for index, value in enumerate(values, start=1):
        text = cell_to_text(value)
        headers.append(text or f"column_{index}")
    return headers


def row_to_text(headers: list[str], values: list[Any]) -> str:
    parts = []
    width = max(len(headers), len(values))
    for index in range(width):
        value = cell_to_text(values[index] if index < len(values) else "")
        if not value:
            continue
        header = headers[index] if index < len(headers) else f"column_{index + 1}"
        parts.append(f"{header}: {value}")
    return "; ".join(parts)


def read_delimited_file(path: Path) -> tuple[list[dict], int, int]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    rows = []
    empty_rows = 0
    with path.open(newline="", encoding="utf-8-sig") as source_file:
        reader = csv.reader(source_file, delimiter=delimiter)
        try:
            header_row = next(reader)
        except StopIteration:
            return [], 0, 0

        headers = normalize_headers(header_row)
        total_rows = 1
        for row_number, row in enumerate(reader, start=2):
            total_rows = row_number
            text = row_to_text(headers, row)
            if text:
                rows.append({"row": row_number, "text": text})
            else:
                empty_rows += 1
    return rows, empty_rows, total_rows


def read_excel_file(path: Path) -> tuple[list[dict], int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading Excel files requires openpyxl. Install requirements.txt first."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows = []
    empty_rows = 0
    total_rows = 0
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            continue

        headers = normalize_headers(list(header_row))
        total_rows += 1
        for row_number, row in enumerate(iterator, start=2):
            total_rows += 1
            text = row_to_text(headers, list(row))
            if text:
                rows.append(
                    {
                        "sheet": worksheet.title,
                        "row": row_number,
                        "text": text,
                    }
                )
            else:
                empty_rows += 1

    workbook.close()
    return rows, empty_rows, total_rows


def chunk_text(text: str, chunk_size: int = 512, overlap: int = 80) -> list[str]:
    if chunk_size <= 0:
        raise ValueError("chunk_size must be greater than zero")
    if overlap < 0 or overlap >= chunk_size:
        raise ValueError("overlap must be at least zero and smaller than chunk_size")

    words = text.split()
    step = chunk_size - overlap
    return [
        " ".join(words[start : start + chunk_size])
        for start in range(0, len(words), step)
    ]


def chunk_id(filename: str, location: int | str, index: int) -> str:
    value = f"{filename}\0{location}\0{index}".encode()
    return hashlib.sha256(value).hexdigest()


def mysql_document_id(
    table: str,
    row_identity: Any,
    *,
    database: str = MYSQL_DATABASE,
    company_id: str | None = None,
    backend: str = "mysql",
) -> str:
    if company_id is None:
        value = f"{backend}\0{database}\0{table}\0{row_identity}".encode()
    else:
        value = (
            f"{backend}\0{company_id}\0{database}\0{table}\0{row_identity}"
        ).encode()
    return hashlib.sha256(value).hexdigest()


def content_hash(document: str) -> str:
    return hashlib.sha256(document.encode()).hexdigest()


def metadata_value(value: Any) -> bool | int | float | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        try:
            text = value.decode("utf-8")
        except UnicodeDecodeError:
            text = value.hex()
    else:
        text = str(value)
    text = text.strip()
    if not text:
        return None
    if len(text) > MYSQL_METADATA_MAX_CHARS:
        return text[:MYSQL_METADATA_MAX_CHARS]
    return text


def normalize_metadata_key(value: str) -> str:
    chars = []
    for char in value:
        chars.append(char if char.isalnum() else "_")
    normalized = "_".join(part for part in "".join(chars).split("_") if part)
    return normalized[:120] or "value"


def parse_json_like(value: Any) -> Any | None:
    if isinstance(value, (dict, list)):
        return value
    text = cell_to_text(value)
    if not text or text[0] not in "[{":
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def normalize_labeled_text(value: str) -> str:
    text = " ".join(value.split())
    return LABELED_TEXT_PATTERN.sub(r"\n\1: ", text).strip()


def extract_labeled_text_metadata(value: str) -> dict:
    matches = list(
        re.finditer(
            r"(?:^|\n)("
            + "|".join(
                re.escape(label)
                for label in sorted(LABELED_TEXT_KEYS, key=len, reverse=True)
            )
            + r"):\s*",
            value,
        )
    )
    metadata = {}
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(value)
        content = value[start:end].strip()
        if not content:
            continue
        key = f"content_{normalize_metadata_key(match.group(1).lower())}"
        safe_value = metadata_value(content)
        if safe_value is not None:
            metadata[key] = safe_value
    return metadata


def iter_json_scalars(value: Any, prefix: str = ""):
    if isinstance(value, dict):
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            yield from iter_json_scalars(child, child_prefix)
    elif isinstance(value, list):
        for index, child in enumerate(value, start=1):
            child_prefix = f"{prefix}.{index}" if prefix else str(index)
            yield from iter_json_scalars(child, child_prefix)
    else:
        safe_value = metadata_value(value)
        if safe_value is not None:
            yield prefix, safe_value


def collect_json_text_values(value: Any) -> list[str]:
    values = []
    if isinstance(value, dict):
        for key, child in value.items():
            key_text = str(key).lower()
            if key_text in JSON_TEXT_KEYS:
                if isinstance(child, (dict, list)):
                    values.extend(
                        str(safe_value)
                        for _, safe_value in iter_json_scalars(child)
                        if isinstance(safe_value, str) and safe_value
                    )
                else:
                    text = cell_to_text(child)
                    if text:
                        values.append(text)
            else:
                values.extend(collect_json_text_values(child))
    elif isinstance(value, list):
        for child in value:
            values.extend(collect_json_text_values(child))
    return values


def flatten_json_as_document(value: Any) -> str:
    parts = []
    for path, safe_value in iter_json_scalars(value):
        if path:
            parts.append(f"{path}: {safe_value}")
        else:
            parts.append(str(safe_value))
    return "; ".join(parts)


def prepare_content_document(value: Any) -> tuple[str, dict]:
    parsed = parse_json_like(value)
    if parsed is None:
        text = cell_to_text(value)
        if not text:
            return "", {"content_format": "text"}
        document = normalize_labeled_text(text)
        metadata = extract_labeled_text_metadata(document)
        metadata["content_format"] = "labeled_text" if metadata else "text"
        return document, metadata

    selected_texts = collect_json_text_values(parsed)
    if selected_texts:
        document = "\n".join(dict.fromkeys(selected_texts))
    else:
        document = flatten_json_as_document(parsed)

    metadata = {"content_format": "json"}
    for index, (path, safe_value) in enumerate(iter_json_scalars(parsed), start=1):
        if index > MYSQL_JSON_METADATA_MAX_FIELDS:
            metadata["content_metadata_truncated"] = True
            break
        key = f"content_{normalize_metadata_key(path)}"
        if key not in metadata:
            metadata[key] = safe_value
    return document, metadata


def mysql_row_identity(row: dict[str, Any], primary_key_column: str | None) -> str:
    if primary_key_column and row.get(primary_key_column) is not None:
        return str(row[primary_key_column])
    normalized = "|".join(
        f"{key}={metadata_value(row[key])}" for key in sorted(row.keys())
    )
    return hashlib.sha256(normalized.encode()).hexdigest()


def prepare_mysql_row(
    row: dict[str, Any],
    content_column: str,
    primary_key_column: str | None,
    *,
    mysql_config: DatabaseRuntimeConfig | None = None,
    company_id: str | None = None,
) -> tuple[str, str, dict] | None:
    document, content_metadata = prepare_content_document(row.get(content_column))
    if not document:
        return None

    database = mysql_config.database if mysql_config else MYSQL_DATABASE
    search_table = mysql_config.search_table if mysql_config else MYSQL_TABLE
    identity = mysql_row_identity(row, primary_key_column)
    metadata = {
        "source_file": database_source_name(mysql_config),
        "source_type": database_backend(mysql_config),
        "source_database": database,
        "source_table": search_table,
        "embedding_model": EMBED_MODEL,
        "source_content_hash": content_hash(document),
    }
    if company_id is not None:
        metadata["company_id"] = company_id
    metadata.update(content_metadata)
    if primary_key_column and row.get(primary_key_column) is not None:
        metadata["primary_key_column"] = primary_key_column
        metadata["primary_key_value"] = metadata_value(row[primary_key_column])

    for column, value in row.items():
        if column == content_column:
            continue
        safe_value = metadata_value(value)
        if safe_value is not None:
            metadata[column] = safe_value

    return (
        mysql_document_id(
            search_table,
            identity,
            database=database,
            company_id=company_id,
            backend=database_backend(mysql_config),
        ),
        document,
        metadata,
    )


def prepare_bm25_index_row(
    row: dict[str, Any],
    content_column: str,
    primary_key_column: str | None,
    *,
    mysql_config: DatabaseRuntimeConfig | None = None,
    company_id: str | None = None,
) -> dict | None:
    content = cell_to_text(row.get(content_column))
    if not content:
        return None

    database = mysql_config.database if mysql_config else MYSQL_DATABASE
    search_table = mysql_config.search_table if mysql_config else MYSQL_TABLE
    search_id_column = (
        mysql_config.search_id_column
        if mysql_config
        else MYSQL_SEARCH_ID_COLUMN
    )
    identity = mysql_row_identity(row, primary_key_column)
    product_id = row.get(search_id_column)
    if product_id is None:
        product_id = identity

    rental_fee = metadata_value(row.get("rental_fee"))
    if not isinstance(rental_fee, (int, float)):
        rental_fee = None

    return {
        "doc_id": mysql_document_id(
            search_table,
            identity,
            database=database,
            company_id=company_id,
            backend=database_backend(mysql_config),
        ),
        "product_id": product_id,
        "content": content,
        "main_category_name": metadata_value(row.get("main_category_name")),
        "subcategory_name": metadata_value(row.get("subcategory_name")),
        "state_name": metadata_value(row.get("state_name")),
        "city_name": metadata_value(row.get("city_name")),
        "locality_name": metadata_value(row.get("locality_name")),
        "rental_duration": metadata_value(row.get("rental_duration")),
        "rental_fee": rental_fee,
        "main_category_id": metadata_value(row.get("main_category_id")),
        "subcategory_id": metadata_value(row.get("subcategory_id")),
        "state_id": metadata_value(row.get("state_id")),
        "city_id": metadata_value(row.get("city_id")),
        "locality_id": metadata_value(row.get("locality_id")),
        # These two fields are intentionally optional. Gainr can add them to
        # ads_search_ready later without blocking today's ingestion.
        "ad_type": metadata_value(
            row.get("type", row.get("ad_type"))
        ),
        "is_rent_negotiable": metadata_value(
            row.get("is_rent_negotiable")
        ),
    }


def prepare_pdf(path: Path) -> tuple[list[str], list[str], list[dict], int, int]:
    pages, empty_pages, page_count = read_pdf(path)
    ids = []
    documents = []
    metadatas = []
    for page in pages:
        for index, text in enumerate(
            chunk_text(page["text"], CHUNK_SIZE, CHUNK_OVERLAP)
        ):
            ids.append(chunk_id(path.name, page["page"], index))
            documents.append(text)
            metadatas.append(
                {
                    "source_file": path.name,
                    "page": page["page"],
                    "chunk_index": index,
                    "embedding_model": EMBED_MODEL,
                }
            )
    return ids, documents, metadatas, empty_pages, page_count


def prepare_table(path: Path) -> tuple[list[str], list[str], list[dict], int, int]:
    if path.suffix.lower() in {".csv", ".tsv"}:
        rows, empty_rows, row_count = read_delimited_file(path)
    else:
        rows, empty_rows, row_count = read_excel_file(path)

    ids = []
    documents = []
    metadatas = []
    for row in rows:
        location = (
            f"{row['sheet']}:{row['row']}" if "sheet" in row else str(row["row"])
        )
        for index, text in enumerate(
            chunk_text(row["text"], CHUNK_SIZE, CHUNK_OVERLAP)
        ):
            ids.append(chunk_id(path.name, location, index))
            documents.append(text)
            metadata = {
                "source_file": path.name,
                "source_type": path.suffix.lower().lstrip("."),
                "row": row["row"],
                "chunk_index": index,
                "embedding_model": EMBED_MODEL,
            }
            if "sheet" in row:
                metadata["sheet"] = row["sheet"]
            metadatas.append(metadata)
    return ids, documents, metadatas, empty_rows, row_count


def prepare_source(path: Path) -> tuple[list[str], list[str], list[dict], int, int]:
    if path.suffix.lower() == ".pdf":
        return prepare_pdf(path)
    if path.suffix.lower() in SUPPORTED_EXTENSIONS:
        return prepare_table(path)
    raise ValueError(f"Unsupported source type: {path.suffix}")


def find_source_files() -> list[Path]:
    return sorted(
        [
            path
            for path in SOURCE_FILES_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
        ],
        key=lambda path: path.name.lower(),
    )
